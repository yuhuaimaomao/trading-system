#!/usr/bin/env python3
"""
早报股票追踪记录器 v2.0

功能:
1. 解析早报推送内容，记录股票到 Excel 和数据库
2. 查询真实股票代码/市值（从数据库）
3. 过滤科创板（688 开头）
4. 提取推荐理由关键词
5. 复盘后更新当天行情数据
6. 次日计算准确率（T 日开盘买，T+1 日开盘卖）
"""

import re
import sqlite3
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

from system.utils.logger import get_task_logger

logger = get_task_logger("track")

from system.config.settings import DATABASE_PATH, PROJECT_ROOT

# 项目路径
BASE_DIR = PROJECT_ROOT
EXCEL_PATH = BASE_DIR / "docs/股票追踪表.xlsx"


class StockTracker:
    """股票追踪记录器"""

    # 推荐理由关键词
    REASON_KEYWORDS = [
        "反包",
        "突破",
        "主力净流入",
        "涨停",
        "龙头",
        "趋势",
        "放量",
        "缩量",
        "支撑",
        "压力",
        "缠论",
        "一买",
        "二买",
        "三买",
        "洗盘",
        "拉升",
        "建仓",
        "主升浪",
        "连板",
        "首板",
        "强势",
        "弱势",
        "震荡",
        "回调",
        "反弹",
    ]

    def __init__(self):
        """初始化"""
        try:
            self.wb = openpyxl.load_workbook(EXCEL_PATH)
            if "股票追踪" in self.wb.sheetnames:
                self.ws = self.wb["股票追踪"]
            else:
                # 文件存在但 sheet 名不匹配 → 使用第一个 sheet 或重命名，绝不覆盖
                logger.warning(
                    f"未找到「股票追踪」sheet，当前 sheets: {self.wb.sheetnames}"
                )
                ws_name = self.wb.sheetnames[0]
                self.ws = self.wb[ws_name]
                self.ws.title = "股票追踪"
                logger.info(f"已将 sheet「{ws_name}」重命名为「股票追踪」")
        except FileNotFoundError:
            logger.warning(f"Excel 文件不存在，创建新文件：{EXCEL_PATH}")
            self._create_excel_sheet(from_scratch=True)

    def _create_excel_sheet(self, from_scratch: bool = True):
        """创建 Excel 工作表。from_scratch=True 创建新文件，False 在已有 wb 上新增 sheet。"""
        if from_scratch:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
        self.ws.title = "股票追踪"

        # 设置表头（23 列）
        headers = [
            "推送日期",
            "股票名称",
            "股票代码",
            "所属板块",
            "预期强度",
            "真实市值 (亿)",
            "推荐理由",
            "开盘表现",
            "收盘表现",
            "日内差",
            "是否涨停",
            "次日开盘表现",
            "次日开盘收益",
            "开盘胜率判定",
            "次日收盘收益",
            "收盘胜率判定",
            "次日高点收益",
            "高点胜率判定",
            "次日低点收益",
            "平均股价收益率",
            "平均股价胜率判定",
            "备注",
            "来源",
        ]

        # 定义边框样式
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 定义表头样式
        header_font = Font(bold=True)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = header_alignment

        # 设置列宽
        column_widths = [
            12,
            12,
            10,
            15,
            10,
            12,
            20,
            12,
            12,
            12,
            10,
            12,
            12,
            10,
            12,
            10,
            12,
            10,
            12,
            15,
            15,
            15,
            8,
        ]
        for col, width in enumerate(column_widths, 1):
            self.ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = width

        self.wb.save(EXCEL_PATH)
        logger.info(f"✅ 创建 Excel 工作表：{EXCEL_PATH}")

    def get_db_connection(self):
        """获取数据库连接"""
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        return conn

    def get_stock_code(self, stock_name: str, push_date: str) -> str:
        """
        从数据库查询股票代码

        如果查不到，说明：
        1. 股票名称错误（AI 推荐错误）
        2. 该股票今日停牌（stock_basic 没有记录）
        3. 科创板（应该在解析时已过滤）
        4. 数据采集失败

        Args:
            stock_name: 股票名称
            push_date: 推送日期

        Returns:
            股票代码，如果找不到返回空字符串
        """
        conn = self.get_db_connection()
        cursor = conn.cursor()

        # 从 stock_basic 表查询（推送日期前最近一个交易日）
        cursor.execute(
            """
            SELECT stock_code FROM stock_basic
            WHERE stock_name = ? AND trade_date <= ?
            ORDER BY trade_date DESC LIMIT 1
        """,
            (stock_name, push_date),
        )

        result = cursor.fetchone()
        conn.close()

        if result:
            return result["stock_code"]

        # 查不到，返回空字符串（不记录该股票）
        return ""

    def get_market_cap(self, stock_code: str, push_date: str) -> float:
        """
        查询 T-1 日的真实市值（单位：亿）

        Args:
            stock_code: 股票代码
            push_date: 推送日期

        Returns:
            市值（亿），找不到返回 0
        """
        conn = self.get_db_connection()
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT total_market_cap FROM stock_basic
            WHERE stock_code = ? AND trade_date <= ?
            ORDER BY trade_date DESC LIMIT 1
        """,
            (stock_code, push_date),
        )

        result = cursor.fetchone()
        conn.close()

        if result and result["total_market_cap"]:
            return result["total_market_cap"] / 100000000  # 转换为亿

        return 0

    def is_kechuangban(self, stock_code: str) -> bool:
        """判断是否为科创板（688 开头）"""
        return stock_code.startswith("688")

    def extract_reason_keywords(self, brief_text: str, stock_name: str) -> str:
        """
        从 AI 推荐文本中提取关键词

        Args:
            brief_text: 早报全文
            stock_name: 股票名称

        Returns:
            关键词字符串，如 "反包 + 主力净流入"
        """
        # 找到股票名所在的段落（前后 200 字）
        pattern = f".{{0,200}}{re.escape(stock_name)}.{{0,200}}"
        match = re.search(pattern, brief_text, re.DOTALL)

        if not match:
            return ""

        context = match.group(0)

        # 提取匹配的关键词
        found_keywords = []
        for keyword in self.REASON_KEYWORDS:
            if keyword in context:
                found_keywords.append(keyword)

        # 去重，限制 3 个关键词
        found_keywords = list(dict.fromkeys(found_keywords))[:3]

        return " + ".join(found_keywords)

    def enrich_stock_pool(self, stock_pool: list) -> list:
        """
        完善股票池数据（补充代码、市值等）

        Args:
            stock_pool: AI 返回的股票池列表
            [{'股票名称': 'xx', '股票代码': 'xx', '所属板块': 'xx', ...}, ...]

        Returns:
            完善后的股票列表（格式同 parse_morning_brief 返回值）
        """
        stocks = []
        push_date = datetime.now().strftime("%Y-%m-%d")

        for stock in stock_pool:
            stock_name = stock["股票名称"]
            stock_code = stock["股票代码"]

            # 如果 AI 没提供代码，查询
            if not stock_code:
                logger.info(f"🔍 AI 未提供代码，查询：{stock_name}")
                stock_code = self.get_stock_code(stock_name, push_date)

            # 验证股票代码
            if not stock_code:
                logger.error(f"❌ 股票'{stock_name}'无代码，跳过")
                continue

            # 过滤科创板
            if self.is_kechuangban(stock_code):
                logger.info(f"⚠️ 跳过科创板：{stock_name} ({stock_code})")
                continue

            # 查询市值（如果 AI 提供了市值，优先使用 AI 的）
            ai_market_cap = stock.get("市值", "")
            if ai_market_cap:
                # AI 提供了市值（如 "500 亿"、"84 亿"）
                try:
                    market_cap_str = ai_market_cap.replace("亿", "").strip()
                    market_cap = float(market_cap_str)
                    logger.info(f"✅ 使用 AI 提供的市值：{stock_name} - {market_cap}亿")
                except (ValueError, TypeError):
                    # 解析失败，查询数据库
                    market_cap = self.get_market_cap(stock_code, push_date)
                    logger.info(
                        f"🔍 AI 市值解析失败，查询数据库：{stock_name} - {market_cap:.1f}亿"
                        if market_cap
                        else "查询失败"
                    )
            else:
                # AI 没提供市值，查询数据库
                market_cap = self.get_market_cap(stock_code, push_date)
                logger.info(
                    f"🔍 AI 未提供市值，查询数据库：{stock_name} - {market_cap:.1f}亿"
                    if market_cap
                    else "查询失败"
                )

            # 转换优先级为星级
            priority = stock.get("优先级", "")
            star_rating = 0
            if priority:
                if "P0" in priority:
                    star_rating = 5
                elif "P1" in priority:
                    star_rating = 4
                elif "P2" in priority:
                    star_rating = 3

            stocks.append(
                {
                    "股票名称": stock_name,
                    "股票代码": stock_code,
                    "所属板块": stock.get("所属板块", ""),
                    "sector_code": stock.get("sector_code", ""),
                    "预期强度": "★" * star_rating
                    if star_rating > 0
                    else stock.get("优先级", ""),
                    "star_rating": star_rating,
                    "真实市值": f"{market_cap:.1f}亿" if market_cap else "",
                    "market_cap": market_cap,
                    "推荐理由": stock.get("推荐理由", ""),
                    "买入条件": stock.get("买入条件", ""),
                    "放弃条件": stock.get("放弃条件", ""),
                    "止损位": stock.get("止损位", ""),
                    "目标位": stock.get("目标位", ""),
                }
            )

            logger.info(
                f"✅ 完善股票数据：{stock_name} ({stock_code}) - {stock['所属板块']}"
            )

        # 去重
        seen = set()
        unique_stocks = []
        for stock in stocks:
            key = stock["股票代码"]
            if key not in seen:
                seen.add(key)
                unique_stocks.append(stock)

        logger.info(f"股票池完善完成：{len(unique_stocks)}只股票 (去重后)")
        return unique_stocks

    def parse_morning_brief(self, brief_text: str) -> list:
        """
        解析早盘简报，提取股票信息

        Args:
            brief_text: 早报推送文本

        Returns:
            股票列表，每个股票包含：
            - 股票名称、股票代码、所属板块、预期强度、市值、推荐理由
        """
        stocks = []
        current_plate = None
        current_stars = 0

        lines = brief_text.split("\n")
        for line in lines:
            # 1. 匹配板块行：🔥 板块名 (强度：★★★★★)
            if "🔥" in line and "强度" in line:
                start = line.find("(")
                if start == -1:
                    start = line.find("（")
                end = line.find(")")
                if end == -1:
                    end = line.find("）")

                if start != -1 and end != -1:
                    plate = line[line.find("🔥") + 1 : start].strip()

                    # 提取星级
                    star_content = line[start : end + 1]
                    stars = (
                        star_content.replace("(", "")
                        .replace(")", "")
                        .replace("（", "")
                        .replace("）", "")
                        .replace("强度", "")
                        .replace(":", "")
                        .replace(":", "")
                        .strip()
                    )
                    current_stars = stars.count("★")

                    current_plate = plate
                    logger.info(f"解析到板块：{current_plate}, 强度：{current_stars}星")
                    continue

            # 2. 匹配股票行：【股票名】（市值 XX 亿）
            if current_plate and "【" in line:
                matches = re.findall(r"【(.+?)】", line)
                for stock_name in matches:
                    stock_name = stock_name.strip()

                    # 过滤无效股票名
                    if stock_name in ["主选股", "备选", "股票名", "锚点"]:
                        continue

                    # 查询股票代码
                    push_date = datetime.now().strftime("%Y-%m-%d")
                    stock_code = self.get_stock_code(stock_name, push_date)

                    # 检查股票代码是否存在（AI 推荐错误则跳过）
                    if not stock_code:
                        logger.error(
                            f"❌ AI 推荐错误：股票名称 '{stock_name}' 不存在或无行情数据"
                        )
                        logger.error("   跳过该股票，不记录到追踪表")
                        logger.error("   可能原因：股票名称错误/今日停牌/数据缺失")
                        continue

                    # 过滤科创板
                    if self.is_kechuangban(stock_code):
                        logger.info(f"⚠️ 跳过科创板：{stock_name} ({stock_code})")
                        continue

                    # 查询真实市值
                    market_cap = self.get_market_cap(stock_code, push_date)

                    # 提取推荐理由
                    reason = self.extract_reason_keywords(brief_text, stock_name)

                    stocks.append(
                        {
                            "股票名称": stock_name,
                            "股票代码": stock_code,
                            "所属板块": current_plate,
                            "预期强度": "★" * current_stars,
                            "star_rating": current_stars,
                            "真实市值": f"{market_cap:.1f}亿" if market_cap else "",
                            "market_cap": market_cap,
                            "推荐理由": reason,
                        }
                    )

                    logger.info(
                        f"解析到股票：{stock_name} ({stock_code}) - {current_plate}"
                    )

        # 去重
        seen = set()
        unique_stocks = []
        for stock in stocks:
            key = stock["股票代码"]
            if key not in seen:
                seen.add(key)
                unique_stocks.append(stock)

        logger.info(f"总共解析到 {len(unique_stocks)} 只股票 (去重后)")
        return unique_stocks

    def record_stocks(
        self, stocks: list, push_date: str, brief_text: str = "", source: str = "早报"
    ):
        """
        批量添加股票到 Excel 和数据库（自动去重：同 push_date + stock_code 只写一次）

        Args:
            stocks: 股票列表
            push_date: 推送日期 (YYYY-MM-DD)
            brief_text: 早报全文（用于提取关键词）
            source: 来源（早报/复盘）
        """
        if not stocks:
            logger.warning("⚠️ 没有股票需要记录")
            return

        # 构建 Excel 已有记录的 (push_date, stock_code) 集合，避免追加重复行
        existing_keys = set()
        for row in range(2, self.ws.max_row + 1):
            rd = self.ws.cell(row=row, column=1).value
            rc = self.ws.cell(row=row, column=3).value
            if rd and rc:
                existing_keys.add(f"{rd}|{rc}")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        conn = self.get_db_connection()
        cursor = conn.cursor()

        excel_count = 0
        db_count = 0

        for stock in stocks:
            code = stock.get("股票代码", "")
            key = f"{push_date}|{code}"

            # Excel：已存在则跳过
            if key not in existing_keys:
                self.ws.append(
                    [
                        push_date,
                        stock.get("股票名称", ""),
                        code,
                        stock.get("所属板块", ""),
                        stock.get("预期强度", ""),
                        stock.get("真实市值", ""),
                        stock.get("推荐理由", ""),
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        source,
                    ]
                )
                row = self.ws.max_row
                for col in range(1, 24):
                    self.ws.cell(row=row, column=col).border = thin_border
                existing_keys.add(key)
                excel_count += 1

            # DB：唯一索引防重复，INSERT OR IGNORE 静默跳过
            try:
                cursor.execute(
                    """
                    INSERT OR IGNORE INTO stock_tracker (
                        push_date, stock_code, stock_name, plate, star_rating,
                        market_cap, reason_keywords, source,
                        sector_code, abandon_condition, stop_loss, target_price
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                    (
                        push_date,
                        code,
                        stock.get("股票名称", ""),
                        stock.get("所属板块", ""),
                        stock.get("star_rating", 0),
                        stock.get("market_cap", 0),
                        stock.get("推荐理由", ""),
                        source,
                        stock.get("sector_code", ""),
                        stock.get("放弃条件", ""),
                        stock.get("止损位", ""),
                        stock.get("目标位", ""),
                    ),
                )
                if cursor.rowcount > 0:
                    db_count += 1
            except Exception as e:
                logger.error(f"⚠️ 数据库写入失败：{stock.get('股票名称')} - {e}")

        self.wb.save(EXCEL_PATH)
        conn.commit()
        conn.close()

        skipped = len(stocks) - db_count
        msg = f"✅ 已记录 {db_count} 只股票到 Excel 和数据库"
        if skipped:
            msg += f"（{skipped} 只已存在，跳过）"
        logger.info(msg)

    def update_daily_data(self, trade_date: str):
        """
        补充当天的行情数据（H/I/J/K 列）

        早报: market data = trade_date,       entry = T open
        复盘: market data = next_trading_day(trade_date), entry = T+1 open
        另含追赶——上一交易日推送的复盘股票，其 entry_date 正好是 trade_date。

        records_to_update / market_data 使用 composite key 以避免同股票跨日覆盖。
        """
        from system.config.trading_calendar import (
            get_next_trading_day,
            get_previous_trading_day,
        )

        logger.info(f"开始补充 {trade_date} 的行情数据...")

        conn = self.get_db_connection()
        cursor = conn.cursor()

        # 1️⃣ 收集所有待处理记录（每条记录独立跟踪 push_date / target_date）
        cursor.execute(
            """
            SELECT stock_code, source, push_date FROM stock_tracker
            WHERE push_date = ?
        """,
            (trade_date,),
        )
        rows = cursor.fetchall()

        prev_day = get_previous_trading_day(trade_date)
        two_days_ago = get_previous_trading_day(prev_day) if prev_day else None
        prev_day_catch_up = []
        two_days_ago_catch_up = []
        if prev_day:
            cursor.execute(
                """
                SELECT stock_code, source, push_date FROM stock_tracker
                WHERE push_date = ? AND source = '复盘'
            """,
                (prev_day,),
            )
            prev_day_catch_up = cursor.fetchall()
        if two_days_ago:
            cursor.execute(
                """
                SELECT stock_code, source, push_date FROM stock_tracker
                WHERE push_date = ? AND source = '复盘'
                  AND t_open IS NULL
            """,
                (two_days_ago,),
            )
            two_days_ago_catch_up = cursor.fetchall()

        if not rows and not prev_day_catch_up and not two_days_ago_catch_up:
            logger.info(f"{trade_date} 没有需要更新的股票")
            conn.close()
            return

        next_trade_date = get_next_trading_day(trade_date)

        # records_to_update: list of {stock_code, push_date, target_date, source}
        records = []
        for r in rows:
            if r["source"] == "复盘":
                records.append(
                    {
                        "stock_code": r["stock_code"],
                        "push_date": trade_date,
                        "target_date": next_trade_date,
                        "source": "复盘",
                    }
                )
            else:
                records.append(
                    {
                        "stock_code": r["stock_code"],
                        "push_date": trade_date,
                        "target_date": trade_date,
                        "source": "早报",
                    }
                )
        # T-1 复盘股：推在昨天，需要今天的数据来填 t_open
        for r in prev_day_catch_up:
            records.append(
                {
                    "stock_code": r["stock_code"],
                    "push_date": prev_day,
                    "target_date": trade_date,
                    "source": "复盘追赶",
                }
            )
        # T-2 复盘股：推在前天，需要昨天的数据来填 t_open（不是今天！）
        for r in two_days_ago_catch_up:
            records.append(
                {
                    "stock_code": r["stock_code"],
                    "push_date": two_days_ago,
                    "target_date": prev_day,
                    "source": "复盘追赶",
                }
            )

        morning_count = sum(1 for r in records if r["source"] == "早报")
        review_count = sum(1 for r in records if r["source"] == "复盘")
        catch_up_total = len(prev_day_catch_up) + len(two_days_ago_catch_up)
        logger.info(
            f"需要更新 {len(records)} 只（早报:{morning_count} 复盘:{review_count}"
            f" 追赶:{catch_up_total}）"
        )

        # 2️⃣ 按 target_date 分组抓取 market data（key = "code|target_date"）
        from collections import defaultdict

        by_target_date = defaultdict(list)
        for r in records:
            by_target_date[r["target_date"]].append(r)

        market_data = {}  # key = "code|target_date"
        all_limit_up_stocks = set()

        for target_date, recs in by_target_date.items():
            codes = list({r["stock_code"] for r in recs})
            if not codes or not target_date:
                continue
            ph = ",".join("?" * len(codes))
            cursor.execute(
                f"""
                SELECT stock_code, open, price, prev_close, change_pct
                FROM stock_basic
                WHERE trade_date = ? AND stock_code IN ({ph})
            """,
                [target_date] + codes,
            )
            for row in cursor.fetchall():
                key = f"{row['stock_code']}|{target_date}"
                market_data[key] = {
                    "open": row["open"],
                    "close": row["price"],
                    "prev_close": row["prev_close"],
                    "change_pct": row["change_pct"],
                }
            cursor.execute(
                f"""
                SELECT stock_code FROM limit_pool
                WHERE trade_date = ? AND (pool_type = '涨停' OR pool_type = 'zt')
                AND stock_code IN ({ph})
            """,
                [target_date] + codes,
            )
            all_limit_up_stocks.update(row[0] for row in cursor.fetchall())
            found = sum(
                1 for r in recs if f"{r['stock_code']}|{target_date}" in market_data
            )
            logger.info(f"  {target_date} → {found}/{len(recs)} 只有数据")

        if not market_data:
            logger.warning(f"⚠️ {trade_date} 没有行情数据（可能尚未收盘）")
            conn.close()
            return

        logger.info(f"✅ 获取到 {len(market_data)} 条行情数据")

        # 3️⃣ 逐条更新（每条记录用自己的 push_date + target_date 查 market_data）
        update_count = 0
        for r in records:
            code = r["stock_code"]
            lookup_key = f"{code}|{r['target_date']}"
            if lookup_key not in market_data:
                continue

            data = market_data[lookup_key]
            if data["open"] and data["prev_close"] and data["prev_close"] > 0:
                t_open_pct = (
                    (data["open"] - data["prev_close"]) / data["prev_close"] * 100
                )
            else:
                t_open_pct = 0

            t_intra_diff = data["change_pct"] - t_open_pct if data["change_pct"] else 0
            is_limit_up = 1 if code in all_limit_up_stocks else 0

            try:
                cursor.execute(
                    """
                    UPDATE stock_tracker
                    SET t_open = ?, t_close = ?, t_prev_close = ?,
                        t_change_pct = ?, t_open_pct = ?, t_intra_diff = ?,
                        is_limit_up = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE push_date = ? AND stock_code = ?
                """,
                    (
                        data["open"],
                        data["close"],
                        data["prev_close"],
                        data["change_pct"],
                        t_open_pct,
                        t_intra_diff,
                        is_limit_up,
                        r["push_date"],
                        code,
                    ),
                )
                if cursor.rowcount > 0:
                    update_count += 1
            except Exception as e:
                logger.error(f"⚠️ 更新失败 {code} ({r['push_date']}): {e}")

        conn.commit()
        conn.close()
        logger.info(f"✅ 已更新 {update_count} 只股票行情数据")

        # ✅ 同步更新 Excel
        self._update_excel_daily_data_v2(
            trade_date, records, market_data, all_limit_up_stocks
        )

    def update_next_day_data(self, yesterday: str, today: str):
        """
        补充昨天的次日表现（L/M/N 列）

        早报: entry = T_open, exit = T+1 数据
        复盘: entry = T+1_open, exit = T+2 数据（偏移一天）
              另含追赶逻辑——push_date 在两交易日前且 source='复盘' 的票，
              其 exit 正好是 today 数据。

        yesterday_stocks / today_data 使用 composite key "code|date"
        以避免同股票跨日重复时覆盖遗漏。
        """
        from system.config.trading_calendar import (
            get_next_trading_day,
            get_previous_trading_day,
        )

        logger.info(f"开始补充 {yesterday} 的次日表现（{today} 数据）...")

        conn = self.get_db_connection()
        cursor = conn.cursor()

        # 1️⃣ 查询昨日推荐股票（含 source）
        cursor.execute(
            """
            SELECT stock_code, source, t_open, t_close FROM stock_tracker
            WHERE push_date = ?
        """,
            (yesterday,),
        )
        yesterday_stocks = {}
        for row in cursor.fetchall():
            key = f"{row['stock_code']}|{yesterday}"
            yesterday_stocks[key] = dict(row)
            yesterday_stocks[key]["_push_date"] = yesterday

        # 追赶：两交易日前推送的复盘股票，其 exit 正好是 today
        two_days_ago = get_previous_trading_day(yesterday)
        if two_days_ago:
            cursor.execute(
                """
                SELECT stock_code, source, t_open, t_close FROM stock_tracker
                WHERE push_date = ? AND source = '复盘'
            """,
                (two_days_ago,),
            )
            for row in cursor.fetchall():
                key = f"{row['stock_code']}|{two_days_ago}"
                yesterday_stocks[key] = dict(row)
                yesterday_stocks[key]["_push_date"] = two_days_ago
                yesterday_stocks[key]["_catch_up"] = True

        if not yesterday_stocks:
            logger.info(f"{yesterday} 没有需要更新的股票")
            conn.close()
            return

        logger.info(f"需要更新 {len(yesterday_stocks)} 只股票（含追赶）")

        # 2️⃣ 拆分早报/复盘/追赶，确定各自的 exit 数据日期
        review_exit_date = get_next_trading_day(today)
        morning_codes = []
        review_codes_today = []  # push_date=yesterday, exit=next_trading_day(today)
        review_codes_catch_up = []  # push_date=two_days_ago, exit=today

        for key, info in yesterday_stocks.items():
            code = info["stock_code"]
            if info.get("_catch_up"):
                review_codes_catch_up.append(code)
                info["_exit_date"] = today
            elif info.get("source") == "复盘":
                review_codes_today.append(code)
                info["_exit_date"] = review_exit_date
            else:
                morning_codes.append(code)
                info["_exit_date"] = today

        # 收集各组的 exit market data（key = "code|exit_date"）
        today_data = {}

        def _fetch_market_data(codes, target_date, label):
            if not codes or not target_date:
                return
            placeholders = ",".join("?" * len(codes))
            cursor.execute(
                f"""
                SELECT stock_code, open, price, high, low, avg_price FROM stock_basic
                WHERE trade_date = ? AND stock_code IN ({placeholders})
            """,
                [target_date] + codes,
            )
            for row in cursor.fetchall():
                key = f"{row['stock_code']}|{target_date}"
                today_data[key] = {
                    "open": row["open"],
                    "close": row["price"],
                    "high": row["high"],
                    "low": row["low"],
                    "avg_price": row["avg_price"] or 0,
                    "_exit_date": target_date,
                }
            found = len([c for c in codes if f"{c}|{target_date}" in today_data])
            logger.info(f"  {label}: {target_date} → {found}/{len(codes)} 只有数据")

        _fetch_market_data(morning_codes, today, "早报")
        _fetch_market_data(review_codes_catch_up, today, "复盘追赶")
        _fetch_market_data(review_codes_today, review_exit_date, "复盘")

        logger.info(f"✅ 获取到 {len(today_data)} 条 exit 数据")

        # 3️⃣ 更新
        update_count = 0
        for key, info in yesterday_stocks.items():
            code = info["stock_code"]
            push_date = info["_push_date"]
            exit_date = info["_exit_date"]

            lookup_key = f"{code}|{exit_date}"
            if lookup_key not in today_data:
                logger.debug(
                    f"⏳ {code} ({push_date}) exit={exit_date} 数据尚不可用，跳过"
                )
                continue

            t1 = today_data[lookup_key]

            if not info["t_open"] or info["t_open"] <= 0:
                logger.warning(f"⚠️ {code} ({push_date}) entry 开盘价无效")
                continue

            t_open = info["t_open"]
            t_close = info.get("t_close", 0)
            t1_open = t1["open"]
            t1_close = t1.get("close", 0)
            t1_high = t1.get("high", 0)
            t1_low = t1.get("low", 0)
            t1_avg_price = t1.get("avg_price", 0)

            # 次日开盘表现（vs entry 日收盘）
            if t_close and t_close > 0:
                t1_open_pct = (t1_open - t_close) / t_close * 100
            else:
                t1_open_pct = 0

            # 最终收益 = (exit_open - entry_open) / entry_open
            if t_open > 0:
                final_return = (t1_open - t_open) / t_open * 100
                t1_close_return = (t1_close - t_open) / t_open * 100
                t1_high_return = (t1_high - t_open) / t_open * 100
                t1_low_return = (t1_low - t_open) / t_open * 100
            else:
                final_return = t1_close_return = t1_high_return = t1_low_return = 0

            # 平均股价收益率（vs entry 日收盘）
            if t_close and t_close > 0:
                avg_price_return = (t1_avg_price - t_close) / t_close * 100
            else:
                avg_price_return = 0

            try:
                cursor.execute(
                    """
                    UPDATE stock_tracker
                    SET t1_open = ?, t1_open_pct = ?, t1_avg_price = ?,
                        final_return = ?,
                        t1_close_return = ?, t1_high_return = ?, t1_low_return = ?,
                        avg_price_return = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE push_date = ? AND stock_code = ?
                """,
                    (
                        t1_open,
                        t1_open_pct,
                        t1_avg_price,
                        final_return,
                        t1_close_return,
                        t1_high_return,
                        t1_low_return,
                        avg_price_return,
                        push_date,
                        code,
                    ),
                )
                if cursor.rowcount > 0:
                    update_count += 1
            except Exception as e:
                logger.error(f"⚠️ 更新失败 {code} ({push_date}): {e}")

        conn.commit()
        conn.close()
        logger.info(f"✅ 已更新 {update_count} 只股票次日表现")

        # ✅ 同步更新 Excel
        self._update_excel_next_day_data(yesterday, yesterday_stocks, today_data)

    def get_statistics(self, start_date: str = None, end_date: str = None) -> dict:
        """
        获取统计数据

        Args:
            start_date: 开始日期（YYYY-MM-DD）
            end_date: 结束日期（YYYY-MM-DD）

        Returns:
            统计字典
        """
        conn = self.get_db_connection()
        cursor = conn.cursor()

        # 基础统计（从收益率列计算胜率）
        query = """
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN final_return > 0 THEN 1 ELSE 0 END) as wins,
                SUM(CASE WHEN final_return < 0 THEN 1 ELSE 0 END) as losses,
                AVG(final_return) as avg_return
            FROM stock_tracker
            WHERE final_return IS NOT NULL
        """

        params = []
        if start_date:
            query += " AND push_date >= ?"
            params.append(start_date)
        if end_date:
            query += " AND push_date <= ?"
            params.append(end_date)

        cursor.execute(query, params)
        result = cursor.fetchone()

        total = result["total"] or 0
        wins = result["wins"] or 0
        losses = result["losses"] or 0
        avg_return = result["avg_return"] or 0

        win_rate = (wins / total * 100) if total > 0 else 0

        stats = {
            "total": total,
            "wins": wins,
            "losses": losses,
            "win_rate": win_rate,
            "avg_return": avg_return,
        }

        # 板块维度统计
        cursor.execute("""
            SELECT
                plate,
                COUNT(*) as total,
                SUM(CASE WHEN final_return > 0 THEN 1 ELSE 0 END) as wins,
                AVG(final_return) as avg_return
            FROM stock_tracker
            WHERE plate IS NOT NULL AND plate != '' AND final_return IS NOT NULL
            GROUP BY plate
            ORDER BY total DESC
        """)

        stats["by_plate"] = [
            {
                "plate": row["plate"],
                "total": row["total"],
                "wins": row["wins"],
                "win_rate": (row["wins"] / row["total"] * 100)
                if row["total"] > 0
                else 0,
                "avg_return": row["avg_return"] or 0,
            }
            for row in cursor.fetchall()
        ]

        # 强度维度统计
        cursor.execute("""
            SELECT
                star_rating,
                COUNT(*) as total,
                SUM(CASE WHEN final_return > 0 THEN 1 ELSE 0 END) as wins,
                AVG(final_return) as avg_return
            FROM stock_tracker
            WHERE star_rating IS NOT NULL AND final_return IS NOT NULL
            GROUP BY star_rating
            ORDER BY star_rating DESC
        """)

        stats["by_star"] = [
            {
                "star_rating": row["star_rating"],
                "total": row["total"],
                "wins": row["wins"],
                "win_rate": (row["wins"] / row["total"] * 100)
                if row["total"] > 0
                else 0,
                "avg_return": row["avg_return"] or 0,
            }
            for row in cursor.fetchall()
        ]

        conn.close()
        return stats

    def _update_excel_daily_data_v2(
        self, trade_date: str, records: list, market_data: dict, limit_up_stocks: set
    ):
        """
        同步更新 Excel 的当日行情数据（H/I/J/K 列）

        records: [{stock_code, push_date, target_date, source}, ...]
        market_data: key = "code|target_date"
        """
        try:
            update_count = 0

            for row in range(2, self.ws.max_row + 1):
                row_date = self.ws.cell(row=row, column=1).value
                row_code = self.ws.cell(row=row, column=3).value

                if not row_date or not row_code:
                    continue

                # 找到匹配的 record（stock_code + push_date）
                record = None
                for r in records:
                    if r["stock_code"] == row_code and r["push_date"] == row_date:
                        record = r
                        break
                if not record:
                    continue

                lookup_key = f"{row_code}|{record['target_date']}"
                if lookup_key not in market_data:
                    continue

                data = market_data[lookup_key]

                if data["open"] and data["prev_close"] and data["prev_close"] > 0:
                    t_open_pct = (
                        (data["open"] - data["prev_close"]) / data["prev_close"] * 100
                    )
                else:
                    t_open_pct = 0

                t_intra_diff = (
                    data["change_pct"] - t_open_pct if data["change_pct"] else 0
                )
                is_limit_up = "是" if row_code in limit_up_stocks else "否"

                self.ws.cell(row=row, column=8, value=f"{t_open_pct:.2f}%")
                self.ws.cell(row=row, column=9, value=f"{data['change_pct']:.2f}%")
                self.ws.cell(row=row, column=10, value=f"{t_intra_diff:.2f}%")
                self.ws.cell(row=row, column=11, value=is_limit_up)

                update_count += 1

            self.wb.save(EXCEL_PATH)
            logger.info(f"✅ Excel 已同步更新 {update_count} 行")

        except Exception as e:
            logger.error(f"⚠️ Excel 更新失败：{e}")

    def _update_excel_daily_data(
        self, trade_date: str, stock_data: dict, limit_up_stocks: set
    ):
        """
        同步更新 Excel 的当日行情数据（H/I/J/K 列）

        stock_data 中每项可含 _source / _effective_date 标记，
        复盘股票的实际交易日期与早报不同，但不影响 H-K 列的计算公式。
        """
        try:
            update_count = 0

            for row in range(2, self.ws.max_row + 1):
                row_date = self.ws.cell(row=row, column=1).value
                row_code = self.ws.cell(row=row, column=3).value

                if row_date != trade_date:
                    continue
                if row_code not in stock_data:
                    continue

                data = stock_data[row_code]

                if data["open"] and data["prev_close"] and data["prev_close"] > 0:
                    t_open_pct = (
                        (data["open"] - data["prev_close"]) / data["prev_close"] * 100
                    )
                else:
                    t_open_pct = 0

                t_intra_diff = (
                    data["change_pct"] - t_open_pct if data["change_pct"] else 0
                )
                is_limit_up = "是" if row_code in limit_up_stocks else "否"

                self.ws.cell(row=row, column=8, value=f"{t_open_pct:.2f}%")  # H
                self.ws.cell(row=row, column=9, value=f"{data['change_pct']:.2f}%")  # I
                self.ws.cell(row=row, column=10, value=f"{t_intra_diff:.2f}%")  # J
                self.ws.cell(row=row, column=11, value=is_limit_up)  # K

                update_count += 1

            self.wb.save(EXCEL_PATH)
            logger.info(f"✅ Excel 已同步更新 {update_count} 行")

        except Exception as e:
            logger.error(f"⚠️ Excel 更新失败：{e}")

    def _update_excel_next_day_data(
        self, yesterday: str, yesterday_stocks: dict, today_data: dict
    ):
        """
        同步更新 Excel 的次日表现数据（L/M/N/O/P/Q/R/S/T/U 列）

        yesterday_stocks / today_data 使用 composite key "code|date"，
        通过 Excel 行的 (row_date, row_code) 和后算出的 exit_date 分别查找。
        """
        try:
            update_count = 0

            for row in range(2, self.ws.max_row + 1):
                row_date = self.ws.cell(row=row, column=1).value
                row_code = self.ws.cell(row=row, column=3).value

                if not row_date or not row_code:
                    continue

                stock_key = f"{row_code}|{row_date}"
                if stock_key not in yesterday_stocks:
                    continue

                info = yesterday_stocks[stock_key]

                if not info.get("t_open") or info["t_open"] <= 0:
                    continue

                exit_date = info.get("_exit_date")
                if not exit_date:
                    continue

                data_key = f"{row_code}|{exit_date}"
                if data_key not in today_data:
                    continue

                t1 = today_data[data_key]

                t_open = info["t_open"]
                t_close = info.get("t_close", 0)
                t1_open = t1.get("open", 0)
                t1_close = t1.get("close", 0)
                t1_high = t1.get("high", 0)
                t1_low = t1.get("low", 0)
                t1_avg_price = t1.get("avg_price", 0)

                if t_close and t_close > 0:
                    t1_open_pct = (t1_open - t_close) / t_close * 100
                else:
                    t1_open_pct = 0

                if t_open > 0:
                    final_return = (t1_open - t_open) / t_open * 100
                    t1_close_return = (t1_close - t_open) / t_open * 100
                    t1_high_return = (t1_high - t_open) / t_open * 100
                    t1_low_return = (t1_low - t_open) / t_open * 100
                else:
                    final_return = t1_close_return = t1_high_return = t1_low_return = 0

                if t_close and t_close > 0:
                    avg_price_return = (t1_avg_price - t_close) / t_close * 100
                else:
                    avg_price_return = 0

                def flag(val):
                    return "✅" if val > 0 else ("❌" if val < 0 else "➖")

                self.ws.cell(row=row, column=12, value=f"{t1_open_pct:.2f}%")  # L
                self.ws.cell(row=row, column=13, value=f"{final_return:.2f}%")  # M
                self.ws.cell(row=row, column=14, value=flag(final_return))  # N
                self.ws.cell(row=row, column=15, value=f"{t1_close_return:.2f}%")  # O
                self.ws.cell(row=row, column=16, value=flag(t1_close_return))  # P
                self.ws.cell(row=row, column=17, value=f"{t1_high_return:.2f}%")  # Q
                self.ws.cell(row=row, column=18, value=flag(t1_high_return))  # R
                self.ws.cell(row=row, column=19, value=f"{t1_low_return:.2f}%")  # S
                self.ws.cell(row=row, column=20, value=f"{avg_price_return:.2f}%")  # T
                self.ws.cell(row=row, column=21, value=flag(avg_price_return))  # U
                self.ws.cell(row=row, column=22, value="")  # V

                update_count += 1

            self.wb.save(EXCEL_PATH)
            logger.info(f"✅ Excel 已同步更新 {update_count} 行次日表现")

        except Exception as e:
            logger.error(f"⚠️ Excel 更新失败：{e}")


if __name__ == "__main__":
    tracker = StockTracker()
    print("✅ StockTracker 初始化成功")
